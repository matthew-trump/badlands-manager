import os
import re
from datetime import datetime
from dotenv import load_dotenv
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials
import requests

load_dotenv()

SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
SERVICE_ACCOUNT_FILE = 'credentials.json'
FOLDER_ID = os.environ['FOLDER_ID']
SHOW_NAME = os.environ['SHOW_NAME']

# Date pattern: m.DD.YY at the start of the filename
DATE_PATTERN = re.compile(r'^(\d{1,2})\.(\d{2})\.(\d{2})\s+')

def parse_filename_date(name: str) -> datetime | None:
    m = DATE_PATTERN.match(name)
    if not m:
        return None
    month, day, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    return datetime(2000 + year, month, day)

def get_next_spreadsheet() -> tuple[str, str, str] | None:
    """Returns (file_id, filename, mime_type) for the next upcoming show."""
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    service = build('drive', 'v3', credentials=creds)

    # List xlsx files in the folder
    query = (
        f"'{FOLDER_ID}' in parents and trashed = false and "
        "(mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
        " or mimeType='application/vnd.google-apps.spreadsheet')"
    )
    results = service.files().list(q=query, fields='files(id, name, mimeType)').execute()
    files = results.get('files', [])

    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    upcoming = []
    past = []
    for f in files:
        date = parse_filename_date(f['name'])
        if not date:
            continue
        entry = (date, f['id'], f['name'], f['mimeType'])
        if date >= today:
            upcoming.append(entry)
        else:
            past.append(entry)

    if upcoming:
        upcoming.sort(key=lambda x: x[0])
        _, file_id, file_name, mime_type = upcoming[0]
        return file_id, file_name, mime_type

    if past:
        past.sort(key=lambda x: x[0], reverse=True)
        _, file_id, file_name, mime_type = past[0]
        print("No upcoming spreadsheets found. Using most recent past file.")
        return file_id, file_name, mime_type

    print("No spreadsheets found.")
    return None

RUNSHEETS_DIR = 'runsheets'
SPONSOR_DOWNLOADS_DIR = 'sponsor-downloads'

GDRIVE_FILE_RE = re.compile(r'/d/([^/]+)')

def download_sponsor_file(url: str, advertiser: str) -> str | None:
    """Download a sponsor file from a public Google Drive URL."""
    m = GDRIVE_FILE_RE.search(url)
    if not m:
        print(f"    Could not parse file ID from URL: {url}")
        return None
    file_id = m.group(1)

    # Resolve filename via a HEAD request to the public download URL
    download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
    session = requests.Session()
    response = session.get(download_url, stream=True)

    # For large files, Drive redirects through a virus-scan confirmation page
    if 'content-disposition' not in response.headers:
        # Look for confirmation token
        for key, value in response.cookies.items():
            if key.startswith('download_warning'):
                response = session.get(download_url, params={'confirm': value}, stream=True)
                break

    content_disposition = response.headers.get('content-disposition', '')
    file_name = None
    if 'filename' in content_disposition:
        # Try quoted filename first, then unquoted
        m = re.search(r'filename="([^"]+)"', content_disposition)
        if not m:
            m = re.search(r'filename=([^;]+)', content_disposition)
        if m:
            file_name = m.group(1).strip()
    if not file_name:
        file_name = f"{advertiser.strip()}.bin"

    os.makedirs(SPONSOR_DOWNLOADS_DIR, exist_ok=True)
    local_path = os.path.join(SPONSOR_DOWNLOADS_DIR, file_name)

    if os.path.exists(local_path):
        print(f"    Already downloaded: {local_path}")
        return local_path

    with open(local_path, 'wb') as f:
        for chunk in response.iter_content(chunk_size=32768):
            f.write(chunk)

    print(f"    Downloaded: {local_path}")
    return local_path

def download_spreadsheet(file_id: str, file_name: str, mime_type: str) -> str:
    """Download the file to the runsheets folder and return the local path."""
    if mime_type == 'application/vnd.google-apps.spreadsheet':
        local_name = file_name if file_name.endswith('.xlsx') else file_name + '.xlsx'
    else:
        local_name = file_name

    local_path = os.path.join(RUNSHEETS_DIR, local_name)

    if os.path.exists(local_path):
        print(f"Already downloaded: {local_path}")
        return local_path

    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    service = build('drive', 'v3', credentials=creds)

    if mime_type == 'application/vnd.google-apps.spreadsheet':
        export_mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        request = service.files().export_media(fileId=file_id, mimeType=export_mime)
    else:
        request = service.files().get_media(fileId=file_id)

    os.makedirs(RUNSHEETS_DIR, exist_ok=True)
    with open(local_path, 'wb') as f:
        f.write(request.execute())

    print(f"Successfully downloaded: {local_path}")
    return local_path

def validate_runsheet(path: str, file_name: str) -> bool:
    """Validate show name and date in the runsheet against env and filename."""
    import pandas as pd
    df = pd.read_excel(path, sheet_name='Run Sheet', header=None)

    sheet_show_name = str(df.iloc[0, 0]).strip()
    if sheet_show_name != SHOW_NAME:
        print(f"SHOW NAME MISMATCH: expected '{SHOW_NAME}', got '{sheet_show_name}'")
        return False
    print(f"Show name confirmed: {sheet_show_name}")

    raw_date = df.iloc[1, 0]
    sheet_date = pd.to_datetime(raw_date).date()
    filename_date = parse_filename_date(file_name).date()
    if sheet_date != filename_date:
        print(f"DATE MISMATCH: filename has {filename_date}, sheet has {sheet_date}")
        return False
    print(f"Date confirmed: {sheet_date}")

    col0 = df.iloc[:, 0]
    sponsors_start = col0[col0.astype(str).str.strip() == "Today's Sponsors"].index[0] + 1
    sponsors = []
    for val in col0.iloc[sponsors_start:]:
        s = str(val).strip() if not pd.isna(val) else ''
        if not s or len(s) > 30:
            break
        sponsors.append(s)
    print("Sponsors:")
    for s in sponsors:
        print(f"  {s}")

    bg_rows = col0[col0.astype(str).str.contains('Background Image', case=False, na=False)]
    if not bg_rows.empty:
        print(f"Background image: {str(bg_rows.iloc[-1]).strip()}")
    else:
        print("Background image: (not found)")

    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Run Sheet']

    # Find the header row by looking for a cell with value 'Ad'
    header_row_idx = None
    for row in ws.iter_rows():
        if row[0].value and str(row[0].value).strip() == 'Ad':
            header_row_idx = row[0].row
            break

    if header_row_idx:
        print("\nAd Schedule:")
        for row in ws.iter_rows(min_row=header_row_idx + 1):
            ad = str(row[0].value).strip() if row[0].value else ''
            if not ad:
                break
            ad_type = row[1].value or ''
            advertiser = row[2].value or ''
            link_cell = row[3]
            url = link_cell.hyperlink.target if link_cell.hyperlink else (link_cell.value or '')
            print(f"  {ad} | {ad_type} | {advertiser}")
            if url:
                download_sponsor_file(url, advertiser)
    else:
        print("Ad schedule: (not found)")

    return True

if __name__ == '__main__':
    result = get_next_spreadsheet()
    if result:
        file_id, file_name, mime_type = result
        print(f"Next show file: {file_name}")
        path = download_spreadsheet(file_id, file_name, mime_type)
        validate_runsheet(path, file_name)